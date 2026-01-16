from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse
import openpyxl

from gestion.models import (
    Estudiante, Asignatura, DetalleInscripcion, Pensum, Planificacion, DocumentoCalificaciones, Seccion, PeriodoAcademico
)
from gestion.utils import apply_excel_styling
from gestion.api.serializers import (
    EstudianteSerializer, AsignaturaSerializer, PensumSerializer,
    PlanificacionSerializer, DocumentoCalificacionesSerializer, CreateUserSerializer, UserSerializer,
    ProgramaSerializer, SeccionSerializer, DocenteSerializer, AdministradorSerializer, PeriodoAcademicoSerializer
)
from gestion.permissions import IsAdmin, IsDocente, IsEstudiante, IsDocenteOrAdminOrOwner, IsDocenteOrAdmin
from django.contrib.auth.models import User



class EstudianteViewSet(viewsets.ModelViewSet):
    queryset = Estudiante.objects.all()
    serializer_class = EstudianteSerializer
    filterset_fields = ['programa', 'cedula']
    search_fields = ['usuario__first_name', 'usuario__last_name', 'usuario__username', 'cedula']

    def get_permissions(self):
        if self.action in ['list']:
            return [IsDocenteOrAdmin()]
        if self.action in ['create', 'destroy']:
            return [IsAdmin()]
        if self.action in ['retrieve', 'progreso', 'descargar_progreso_academico']:
            return [IsDocenteOrAdminOrOwner()]
        if self.action in ['mis_inscripciones', 'mi_info', 'descargar_progreso_academico']:
            return [IsEstudiante()]
        return []

    @action(detail=False, methods=['get'], url_path='descargar-progreso-academico')
    def descargar_progreso_academico(self, request):
        """Genera un archivo Excel con el progreso académico del estudiante."""
        from django.http import HttpResponse
        try:
            import openpyxl
        except ImportError:
            return Response({'error': 'Librería openpyxl no instalada.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            estudiante = Estudiante.objects.get(usuario=request.user)
        except Estudiante.DoesNotExist:
            return Response({'error': 'No tienes perfil de estudiante.'}, status=status.HTTP_404_NOT_FOUND)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Progreso - {estudiante.usuario.first_name}"

        from openpyxl.styles import Font
        header_font = Font(bold=True)

        ws['A1'] = "REPORTE DE PROGRESO ACADÉMICO"
        ws['A1'].font = header_font
        ws.merge_cells('A1:F1')

        ws['A2'] = "Estudiante:"
        ws['B2'] = estudiante.usuario.get_full_name()

        ws['A3'] = "Cédula:"
        ws['B3'] = estudiante.cedula
        
        headers = ['Código', 'Asignatura', 'Semestre', 'Créditos', 'Nota', 'Estatus']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.font = header_font

        detalles = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante
        ).select_related('asignatura').order_by('asignatura__semestre', 'asignatura__codigo')

        row_num = 6
        for det in detalles:
            ws.cell(row=row_num, column=1, value=det.asignatura.codigo)
            ws.cell(row=row_num, column=2, value=det.asignatura.nombre_asignatura)
            ws.cell(row=row_num, column=3, value=det.asignatura.semestre)
            ws.cell(row=row_num, column=4, value=det.asignatura.creditos)
            ws.cell(row=row_num, column=5, value=float(det.nota_final) if det.nota_final is not None else "N/A")
            ws.cell(row=row_num, column=6, value=det.estatus)
            row_num += 1

        apply_excel_styling(ws, 5, custom_widths={'A': 13.0, 'B': 42.0})

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Record_Academico_{estudiante.cedula}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='mi-horario')
    def mi_horario(self, request):
        """Devuelve el horario de clases del estudiante."""
        try:
            estudiante = Estudiante.objects.get(usuario=request.user)
        except Estudiante.DoesNotExist:
            return Response({'error': 'No eres un estudiante.'}, status=status.HTTP_403_FORBIDDEN)

        periodo_actual = PeriodoAcademico.objects.filter(activo=True).first()
        if not periodo_actual:
            return Response({'error': 'No hay período académico activo.'}, status=status.HTTP_404_NOT_FOUND)

        inscripciones = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante,
            inscripcion__periodo=periodo_actual
        ).select_related('seccion', 'asignatura', 'seccion__docente')

        horario_data = []
        colores = [
            {'bg': 'bg-blue-100 dark:bg-blue-900/30', 'border': 'border-blue-200 dark:border-blue-800', 'text': 'text-blue-800 dark:text-blue-200'},
            {'bg': 'bg-green-100 dark:bg-green-900/30', 'border': 'border-green-200 dark:border-green-800', 'text': 'text-green-800 dark:text-green-200'},
            {'bg': 'bg-purple-100 dark:bg-purple-900/30', 'border': 'border-purple-200 dark:border-purple-800', 'text': 'text-purple-800 dark:text-purple-200'},
            {'bg': 'bg-amber-100 dark:bg-amber-900/30', 'border': 'border-amber-200 dark:border-amber-800', 'text': 'text-amber-800 dark:text-amber-200'},
            {'bg': 'bg-rose-100 dark:bg-rose-900/30', 'border': 'border-rose-200 dark:border-rose-800', 'text': 'text-rose-800 dark:text-rose-200'},
            {'bg': 'bg-indigo-100 dark:bg-indigo-900/30', 'border': 'border-indigo-200 dark:border-indigo-800', 'text': 'text-indigo-800 dark:text-indigo-200'},
            {'bg': 'bg-teal-100 dark:bg-teal-900/30', 'border': 'border-teal-200 dark:border-teal-800', 'text': 'text-teal-800 dark:text-teal-200'},
            {'bg': 'bg-cyan-100 dark:bg-cyan-900/30', 'border': 'border-cyan-200 dark:border-cyan-800', 'text': 'text-cyan-800 dark:text-cyan-200'},
        ]
        
        asignatura_color_map = {}
        color_index = 0

        for detalle in inscripciones:
            if not detalle.seccion:
                continue
            
            if detalle.asignatura.codigo not in asignatura_color_map:
                asignatura_color_map[detalle.asignatura.codigo] = colores[color_index % len(colores)]
                color_index += 1
            
            estilos = asignatura_color_map[detalle.asignatura.codigo]

            horarios = detalle.seccion.horarios.all()
            for h in horarios:
                horario_data.append({
                    'id': h.id,
                    'dia': h.dia,
                    'hora_inicio': h.hora_inicio.strftime('%H:%M'),
                    'hora_fin': h.hora_fin.strftime('%H:%M'),
                    'asignatura': detalle.asignatura.nombre_asignatura,
                    'codigo': detalle.asignatura.codigo,
                    'seccion': detalle.seccion.codigo_seccion,
                    'aula': h.aula,
                    'docente': detalle.seccion.docente.get_full_name() if detalle.seccion.docente else 'Sin asignar',
                })

            if not horarios.exists():
                horario_data.append({
                    'id': f'placeholder-{detalle.id}',
                    'dia': 0, # 0 para que no aparezca en la grilla regular (1-6)
                    'hora_inicio': '00:00',
                    'hora_fin': '00:00',
                    'asignatura': detalle.asignatura.nombre_asignatura,
                    'codigo': detalle.asignatura.codigo,
                    'seccion': detalle.seccion.codigo_seccion,
                    'aula': 'N/A',
                    'docente': detalle.seccion.docente.get_full_name() if detalle.seccion.docente else 'Sin asignar',
                    'estilos': estilos
                })
        
        return Response({
            'estudiante': estudiante.usuario.get_full_name(),
            'cedula': estudiante.cedula,
            'periodo': periodo_actual.nombre_periodo,
            'carrera': estudiante.programa.nombre_programa if estudiante.programa else '',
            'horario': horario_data
        })

    @action(detail=False, methods=['get'], url_path='descargar-horario')
    def descargar_horario(self, request):
        """Genera y descarga el horario en formato Excel (Diseño Exacto Imagen)."""
        try:
            estudiante = Estudiante.objects.get(usuario=request.user)
        except Estudiante.DoesNotExist:
            return Response({'error': 'No eres un estudiante.'}, status=status.HTTP_403_FORBIDDEN)

        periodo_actual = PeriodoAcademico.objects.filter(activo=True).first()
        periodo_str = periodo_actual.nombre_periodo if periodo_actual else "N/A"

        inscripciones = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante,
            inscripcion__periodo=periodo_actual
        ).select_related('seccion', 'asignatura', 'seccion__docente').prefetch_related('seccion__horarios')

        grid_data = {day: {} for day in range(1, 7)}
        
        subjects_list = {}
        
        bloques_data = [
            (1, '07:00 - 07:45'), (2, '07:45 - 08:30'), (3, '08:30 - 09:15'),
            (4, '09:15 - 10:00'), (5, '10:00 - 10:45'), (6, '10:45 - 11:30'),
            (7, '11:30 - 12:15'), (8, '12:15 - 13:00'), (9, '13:00 - 13:45'),
            (10, '13:45 - 14:30'), (11, '14:30 - 15:15'), (12, '15:15 - 16:00'),
            (13, '16:00 - 16:45'), (14, '16:45 - 17:30')
        ]

        start_to_block = { label.split(' - ')[0]: bid for bid, label in bloques_data }

        for det in inscripciones:
            if not det.seccion: continue
            
            uid = f"{det.asignatura.codigo}-{det.seccion.codigo_seccion}"
            if uid not in subjects_list:
                subjects_list[uid] = {
                    'codigo': det.asignatura.codigo,
                    'asignatura': det.asignatura.nombre_asignatura,
                    'semestre': det.asignatura.semestre,
                    'seccion': det.seccion.codigo_seccion,
                    'docente': det.seccion.docente.get_full_name() if det.seccion.docente else "SIN ASIGNAR"
                }

            for h in det.seccion.horarios.all():
                start_str = h.hora_inicio.strftime('%H:%M')
                
                start_block = start_to_block.get(start_str)
                if not start_block:
                     for bid, label in bloques_data:
                        if label.startswith(start_str):
                            start_block = bid
                            break
                
                if start_block:
                    h_total = (h.hora_fin.hour * 60 + h.hora_fin.minute) - (h.hora_inicio.hour * 60 + h.hora_inicio.minute)
                    num_blocks = max(1, round(h_total / 45))
                    
                    room_info = f"({h.aula})" if h.aula else "(N/A)"
                    cell_text = f"{det.asignatura.nombre_asignatura}\n{room_info}"

                    for i in range(num_blocks):
                        current_block = start_block + i
                        if current_block > 14: break 
                        grid_data[h.dia][current_block] = {'text': cell_text, 'uid': uid}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Horario {estudiante.cedula}"
        ws.sheet_view.showGridLines = False

        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        COLOR_HEADER_BG = "4472C4" # Azul Oscuro (Títulos)
        COLOR_HEADER_TXT = "FFFFFF"
        COLOR_GRID_HEADER_BG = "4472C4" 
        COLOR_CLASS_BG = "CCC0DA" # Lila/Morado claro (Celdas de clase)
        COLOR_BORDER = "000000"

        font_title = Font(name='Calibri', size=11, bold=True, color=COLOR_HEADER_TXT)
        font_header_row = Font(name='Calibri', size=11, bold=True, color=COLOR_HEADER_TXT) # Grid headers
        font_cell = Font(name='Calibri', size=10)
        font_cell_bold = Font(name='Calibri', size=10, bold=True) # Bloque numbers
        
        fill_header = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        fill_class = PatternFill("solid", fgColor=COLOR_CLASS_BG)
        
        border_thin = Side(border_style="thin", color=COLOR_BORDER)
        border_all = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws['A1'] = "HORARIO DE CLASES"
        ws['A1'].font = font_title
        ws['A1'].fill = fill_header
        ws['A1'].alignment = align_center
        ws.merge_cells('A1:H1')

        ws['A2'] = f"Estudiante: {estudiante.usuario.get_full_name()}"
        ws['A2'].alignment = align_left
        ws['A2'].border = border_all
        ws.merge_cells('A2:D2') # A,B,C,D
        
        ws['E2'] = f"Periodo: {periodo_str}"
        ws['E2'].alignment = align_left
        ws['E2'].border = border_all
        ws.merge_cells('E2:H2') # E,F,G,H

        ws['A3'] = f"Cédula: {estudiante.cedula}"
        ws['A3'].alignment = align_left
        ws['A3'].border = border_all
        ws.merge_cells('A3:D3')

        ws['E3'] = f"Carrera: {estudiante.programa.nombre_programa if estudiante.programa else ''}"
        ws['E3'].alignment = align_left
        ws['E3'].border = border_all
        ws.merge_cells('E3:H3')
        

        headers = ["BLOQUE", "HORAS", "LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO"]
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=text)
            cell.font = font_header_row
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_all

        start_row = 6
        
        for i, (bid, time_label) in enumerate(bloques_data):
            row = start_row + i
            c1 = ws.cell(row=row, column=1, value=bid)
            c1.font = font_cell_bold
            c1.alignment = align_center
            c1.border = border_all
            
            c2 = ws.cell(row=row, column=2, value=time_label)
            c2.font = font_cell
            c2.alignment = align_center
            c2.border = border_all

        for d_idx in range(1, 7): # Dias 1 a 6
            col_idx = 2 + d_idx # Lunes es col 3 (C)
            current_uid = None
            merge_start_row = None
            
            for b_idx in range(1, 15):
                current_row = start_row + (b_idx - 1)
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border_all # Default border empty cells
                
                data = grid_data[d_idx].get(b_idx)
                
                if data:
                    new_uid = data['uid']
                    text = data['text']
                    
                    if new_uid == current_uid:
                        pass
                    else:
                        if current_uid is not None:
                            if merge_start_row < current_row - 1:
                                ws.merge_cells(start_row=merge_start_row, start_column=col_idx, end_row=current_row-1, end_column=col_idx)

                        cell.value = text
                        cell.font = font_cell_bold
                        cell.fill = fill_class
                        cell.alignment = align_center
                        current_uid = new_uid
                        merge_start_row = current_row
                        
                    cell.fill = fill_class
                else:
                    if current_uid is not None:
                         if merge_start_row < current_row: # end was prev row
                             ws.merge_cells(start_row=merge_start_row, start_column=col_idx, end_row=current_row-1, end_column=col_idx)
                         current_uid = None
                         merge_start_row = None
            
            if current_uid is not None:
                 if merge_start_row < (start_row + 14):
                     ws.merge_cells(start_row=merge_start_row, start_column=col_idx, end_row=start_row+13, end_column=col_idx)

        det_head_row = 21
        
        
        ws.column_dimensions['A'].width = 9.50   # N
        ws.column_dimensions['B'].width = 13.50  # Codigo / Horas
        ws.column_dimensions['C'].width = 18  # Lunes / Asig pt1
        ws.column_dimensions['D'].width = 18  # Martes / Asig pt2
        ws.column_dimensions['E'].width = 18  # Miercoles / Semestre
        ws.column_dimensions['F'].width = 18  # Jueves / Sección
        ws.column_dimensions['G'].width = 18  # Viernes / Docente pt1
        ws.column_dimensions['H'].width = 18  # Sabado / Docente pt2

        c = ws.cell(row=det_head_row, column=1, value="N°")
        c.fill = fill_header; c.font = font_header_row; c.alignment = align_center; c.border = border_all
        
        c = ws.cell(row=det_head_row, column=2, value="CÓDIGO")
        c.fill = fill_header; c.font = font_header_row; c.alignment = align_center; c.border = border_all
        
        c = ws.cell(row=det_head_row, column=3, value="ASIGNATURA")
        c.fill = fill_header; c.font = font_header_row; c.alignment = align_center; c.border = border_all
        ws.merge_cells(start_row=det_head_row, start_column=3, end_row=det_head_row, end_column=4)
        ws.cell(row=det_head_row, column=4).border = border_all

        c = ws.cell(row=det_head_row, column=5, value="SEMESTRE")
        c.fill = fill_header; c.font = font_header_row; c.alignment = align_center; c.border = border_all
        
        c = ws.cell(row=det_head_row, column=6, value="SECCIÓN")
        c.fill = fill_header; c.font = font_header_row; c.alignment = align_center; c.border = border_all
        
        c = ws.cell(row=det_head_row, column=7, value="DOCENTE")
        c.fill = fill_header; c.font = font_header_row; c.alignment = align_center; c.border = border_all
        ws.merge_cells(start_row=det_head_row, start_column=7, end_row=det_head_row, end_column=8)
        ws.cell(row=det_head_row, column=8).border = border_all

        sorted_subs = sorted(subjects_list.values(), key=lambda x: (x['semestre'], x['codigo']))
        row_idx = det_head_row + 1
        
        for idx, sub in enumerate(sorted_subs, 1):
            ws.cell(row=row_idx, column=1, value=idx).border = border_all
            ws.cell(row=row_idx, column=1).alignment = align_center
            
            ws.cell(row=row_idx, column=2, value=sub['codigo']).border = border_all
            ws.cell(row=row_idx, column=2).alignment = align_center
            
            ws.cell(row=row_idx, column=3, value=sub['asignatura']).border = border_all
            ws.cell(row=row_idx, column=3).alignment = align_center
            ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=4)
            ws.cell(row=row_idx, column=4).border = border_all
            
            ws.cell(row=row_idx, column=5, value=sub['semestre']).border = border_all
            ws.cell(row=row_idx, column=5).alignment = align_center
            
            ws.cell(row=row_idx, column=6, value=sub['seccion']).border = border_all
            ws.cell(row=row_idx, column=6).alignment = align_center
            
            ws.cell(row=row_idx, column=7, value=sub['docente']).border = border_all
            ws.cell(row=row_idx, column=7).alignment = align_center
            ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=8)
            ws.cell(row=row_idx, column=8).border = border_all
            
            row_idx += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Horario_{estudiante.cedula}.xlsx"'
        wb.save(response)
        return response

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Horario de Clases"

        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        font_bold = Font(bold=True)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Azul similar al frontend
        font_white = Font(color="FFFFFF", bold=True)

        ws['A1'] = "HORARIO DE CLASES"
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = align_center

        ws['A2'] = f"Estudiante: {estudiante.usuario.get_full_name()}"
        ws['A3'] = f"Cédula: {estudiante.cedula}"
        ws['D2'] = f"Periodo: {periodo_actual.nombre_periodo}"
        ws['D3'] = f"Carrera: {estudiante.programa.nombre_programa if estudiante.programa else ''}"

        ws.column_dimensions['A'].width = 8   # Bloque
        ws.column_dimensions['B'].width = 15  # Hora
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 25 # Días

        headers = ["BLOQUE", "HORAS", "LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO"]
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=text)
            cell.font = font_white
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin

        bloques = [
            (1, "07:00", "07:45"), (2, "07:45", "08:30"), (3, "08:30", "09:15"), (4, "09:15", "10:00"),
            (5, "10:00", "10:45"), (6, "10:45", "11:30"), (7, "11:30", "12:15"), (8, "12:15", "13:00"),
            (9, "13:00", "13:45"), (10, "13:45", "14:30"), (11, "14:30", "15:15"), (12, "15:15", "16:00"),
            (13, "16:00", "16:45"), (14, "16:45", "17:30")
        ]
        
        dia_col_map = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8}

        current_row = 6
        for b_id, b_start, b_end in bloques:
            c_id = ws.cell(row=current_row, column=1, value=b_id)
            c_id.alignment = align_center
            c_id.border = border_thin
            c_id.font = font_bold
            
            c_time = ws.cell(row=current_row, column=2, value=f"{b_start} - {b_end}")
            c_time.alignment = align_center
            c_time.border = border_thin

            for dia_iso in range(1, 7):
                col_idx = dia_col_map[dia_iso]
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border_thin
                cell.alignment = align_center
                
                
                clase_info = schedule_map.get((dia_iso, b_start))
                if clase_info:
                    cell.value = f"{clase_info['asignatura']}\n({clase_info['aula']})"
                else:
                    cell.value = ""

            current_row += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Horario_{estudiante.cedula}.xlsx"'
        wb.save(response)
        return response

    @action(detail=True, methods=['get'])
    def progreso(self, request, pk=None):
        estudiante = self.get_object()
        avance = estudiante.calcular_avance()

        total_asignaturas = Estudiante.objects.filter(pk=estudiante.pk).exists() and estudiante.programa and estudiante.programa and estudiante.programa.asignatura_set.count() or 0
        aprobadas = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante,
            nota_final__gte=10
        ).count()

        nombre = ''
        try:
            nombre = estudiante.usuario.get_full_name()
        except Exception:
            nombre = str(estudiante.usuario)

        datos = {
            'id': estudiante.id,
            'cedula': estudiante.cedula,
            'nombre': nombre,
            'programa': estudiante.programa.nombre_programa if estudiante.programa else '',
            'porcentaje_avance': avance,
            'total_asignaturas': total_asignaturas,
            'aprobadas': aprobadas,
        }
        return Response(datos)

    @action(detail=False, methods=['get'], url_path='mis-inscripciones')
    def mis_inscripciones(self, request):
        """Devuelve el estado de inscripción de cada asignatura para el estudiante autenticado."""
        user = request.user
        
        try:
            estudiante = Estudiante.objects.get(usuario=user)
        except Estudiante.DoesNotExist:
            return Response({'error': 'No tienes perfil de estudiante.'}, status=status.HTTP_400_BAD_REQUEST)
        
        detalles = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante
        ).select_related('asignatura', 'seccion')
        
        inscripciones_map = {}
        for detalle in detalles:
            codigo = detalle.asignatura.codigo
            if codigo not in inscripciones_map or detalle.id > inscripciones_map[codigo]['id']:
                inscripciones_map[codigo] = {
                    'id': detalle.id,
                    'asignatura_id': detalle.asignatura.id,
                    'codigo': codigo,
                    'estatus': detalle.estatus,
                    'nota_final': float(detalle.nota_final) if detalle.nota_final else None,
                    'seccion_id': detalle.seccion_id,
                    'seccion_codigo': detalle.seccion.codigo_seccion if detalle.seccion else None,
                }
        
        return Response({
            'estudiante_id': estudiante.id,
            'programa': estudiante.programa.nombre_programa if estudiante.programa else '',
            'inscripciones': inscripciones_map
        })

    @action(detail=False, methods=['get'], url_path='mi-info')
    def mi_info(self, request):
        """Devuelve información de progreso del estudiante autenticado."""
        user = request.user
        
        try:
            estudiante = Estudiante.objects.get(usuario=user)
        except Estudiante.DoesNotExist:
            return Response({'error': 'No tienes perfil de estudiante.'}, status=status.HTTP_400_BAD_REQUEST)
        
        avance = estudiante.calcular_avance()
        
        total_asignaturas = estudiante.programa.asignatura_set.count() if estudiante.programa else 0
        aprobadas = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante,
            nota_final__gte=10
        ).count()

        nombre = ''
        try:
            nombre = estudiante.usuario.get_full_name()
        except Exception:
            nombre = str(estudiante.usuario)

        return Response({
            'id': estudiante.id,
            'cedula': estudiante.cedula,
            'nombre': nombre,
            'nombre_completo': nombre,
            'programa': estudiante.programa.nombre_programa if estudiante.programa else '',
            'porcentaje_avance': avance,
            'total_asignaturas': total_asignaturas,
            'aprobadas': aprobadas,
            'uc_actuales': estudiante.get_uc_periodo_actual() if hasattr(estudiante, 'get_uc_periodo_actual') else 0
        })


    @action(detail=False, methods=['get'])
    def reporte_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listado Estudiantes"
        
        from openpyxl.styles import Font
        header_font = Font(bold=True)
        title_font = Font(bold=True, size=14)

        ws['A1'] = "LISTADO GENERAL DE ESTUDIANTES"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')

        headers = ['Nombres', 'Apellidos', 'Cédula', 'Teléfono', 'Correo', 'Carrera', 'Avance (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font

        row_num = 4
        for est in Estudiante.objects.select_related('usuario', 'programa').all():
            ws.cell(row=row_num, column=1, value=est.usuario.first_name)
            ws.cell(row=row_num, column=2, value=est.usuario.last_name)
            ws.cell(row=row_num, column=3, value=est.cedula)
            ws.cell(row=row_num, column=4, value=est.telefono)
            ws.cell(row=row_num, column=5, value=est.usuario.email)
            ws.cell(row=row_num, column=6, value=est.programa.nombre_programa if est.programa else '')
            ws.cell(row=row_num, column=7, value=f"{est.calcular_avance()}")
            row_num += 1

        apply_excel_styling(ws, 3)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Listado_Estudiantes.xlsx'
        wb.save(response)
        return response


class AsignaturaViewSet(viewsets.ModelViewSet):
    queryset = Asignatura.objects.all()
    serializer_class = AsignaturaSerializer
    filterset_fields = ['programa', 'semestre']
    search_fields = ['nombre_asignatura', 'codigo']

    @action(detail=True, methods=['post'], url_path='assign-docente')
    def assign_docente(self, request, pk=None):
        asignatura = self.get_object()
        codigo_seccion = request.data.get('codigo_seccion')
        docente_id = request.data.get('docente') # User ID
        
        horario_data = request.data.get('horario') # { dia: 1, bloque_inicio: 1, bloque_fin: 2 }

        if not codigo_seccion:
            return Response({'error': 'Código de sección requerido'}, status=400)
        
        docente_user = None
        if docente_id:
            from gestion.models import Docente
            try:
                docente_obj = Docente.objects.get(pk=docente_id)
                docente_user = docente_obj.usuario
            except Docente.DoesNotExist:
                return Response({'error': 'Docente no encontrado'}, status=404)

        if not docente_user:
            Seccion.objects.filter(asignatura=asignatura, codigo_seccion=codigo_seccion).delete()
            return Response({'status': 'deleted'})

        seccion, created = Seccion.objects.update_or_create(
            asignatura=asignatura,
            codigo_seccion=codigo_seccion,
            defaults={'docente': docente_user}
        )
        
        
        if horario_data:
            dia = int(horario_data.get('dia'))
            bloque_inicio = int(horario_data.get('bloque_inicio')) # 1-14
            bloque_fin = int(horario_data.get('bloque_fin')) # 1-14
            
            if dia and bloque_inicio and bloque_fin:
                from gestion.models import Horario
                from datetime import time
                
                BLOCK_MAP = {
                    1: "07:00", 2: "07:45", 3: "08:30", 4: "09:15", 
                    5: "10:00", 6: "10:45", 7: "11:30", 8: "12:15", 
                    9: "13:00", 10: "13:45", 11: "14:30", 12: "15:15", 
                    13: "16:00", 14: "16:45"
                }
                
                BLOCK_END_MAP = {
                    1: "07:45", 2: "08:30", 3: "09:15", 4: "10:00", 
                    5: "10:45", 6: "11:30", 7: "12:15", 8: "13:00", 
                    9: "13:45", 10: "14:30", 11: "15:15", 12: "16:00", 
                    13: "16:45", 14: "17:30"
                }
                
                t_start_str = BLOCK_MAP.get(bloque_inicio)
                t_end_str = BLOCK_END_MAP.get(bloque_fin)
                
                aula = horario_data.get('aula', '')

                if t_start_str and t_end_str:
                     h_start = int(t_start_str.split(':')[0])
                     m_start = int(t_start_str.split(':')[1])
                     h_end = int(t_end_str.split(':')[0])
                     m_end = int(t_end_str.split(':')[1])
                     
                     start_time = time(h_start, m_start)
                     end_time = time(h_end, m_end)
                     
                     Horario.objects.filter(seccion=seccion).delete()
                     
                     Horario.objects.update_or_create(
                         seccion=seccion,
                         dia=dia,
                         defaults={
                             'hora_inicio': start_time,
                             'hora_fin': end_time,
                             'aula': aula
                         }
                     )

        try:
            from gestion.models import Docente
            from gestion.notifications import notify_docente_assignment
            if docente_user and hasattr(docente_user, 'docente'):
                 notify_docente_assignment(docente_user.docente, seccion)
        except Exception as e:
            print(f"Error enviando correo al docente: {e}")
            
        return Response(SeccionSerializer(seccion).data)

    @action(detail=True, methods=['post'], url_path='assign-tutor')
    def assign_tutor(self, request, pk=None):
        asignatura = self.get_object()
        docente_id = request.data.get('docente')
        tutor_type = request.data.get('type', 'generic') # generic, academic, community

        if not docente_id:
            return Response({'error': 'Docente ID requerido'}, status=400)
        
        from gestion.models import Docente
        try:
            docente_obj = Docente.objects.get(pk=docente_id)
            docente_user = docente_obj.usuario
        except Docente.DoesNotExist:
            return Response({'error': 'Docente no encontrado'}, status=404)

        if asignatura.tutores.count() >= 10:
             return Response({'error': "Límite de tutores alcanzado (10)"}, status=400)

        asignatura.tutores.add(docente_user)
        return Response({'status': 'Tutor assigned'})



    @action(detail=True, methods=['post'], url_path='remove-tutor')
    def remove_tutor(self, request, pk=None):
        asignatura = self.get_object()
        docente_id = request.data.get('docente')
        tutor_type = request.data.get('type', 'generic')

        if not docente_id:
            return Response({'error': 'Docente ID requerido'}, status=400)
        
        try:
            docente = User.objects.get(pk=docente_id)
        except User.DoesNotExist:
            return Response({'error': 'Docente no encontrado'}, status=404)

        asignatura.tutores.remove(docente)
        return Response({'status': 'Tutor removed'})


class ProgramaViewSet(viewsets.ModelViewSet):
    from gestion.models import Programa
    queryset = Programa.objects.all()
    serializer_class = ProgramaSerializer
    permission_classes = []
    
    def get_permissions(self):
        from rest_framework.permissions import IsAuthenticated
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]
        return [IsAdmin()]

class UserManagementViewSet(viewsets.ViewSet):
    """Endpoints para que el administrador cree usuarios (docente/estudiante)."""
    permission_classes = [IsAdmin]

    def create(self, request):
        serializer = CreateUserSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)


from rest_framework.views import APIView
from gestion.models import UserActivity
from django.utils import timezone
from datetime import timedelta

class OnlineUsersView(APIView):
    """Devuelve usuarios activos en los últimos 5 minutos."""

    def get(self, request):
        time_threshold = timezone.now() - timedelta(minutes=5)
        active_users = UserActivity.objects.filter(last_activity__gte=time_threshold).select_related('user')
        
        data = []
        for activity in active_users:
            user = activity.user
            
            role = 'Desconocido'
            if user.is_superuser or user.is_staff or user.groups.filter(name='Administrador').exists():
                role = 'Administrador'
            elif hasattr(user, 'docente'):
                role = 'Docente'
            elif hasattr(user, 'estudiante'):
                role = 'Estudiante'

            name = user.get_full_name() or user.username
            
            student_id = user.estudiante.id if hasattr(user, 'estudiante') else None
            
            data.append({
                'id': user.id, # User ID
                'student_id': student_id,
                'username': user.username,
                'name': name,
                'email': user.email,
                'role': role,
                'status': 'online',
                'last_activity': activity.last_activity,
                'device': activity.device_type
            })
            
        return Response(data)


class DocenteViewSet(viewsets.ModelViewSet):
    from gestion.models import Docente
    queryset = Docente.objects.all()
    serializer_class = DocenteSerializer
    permission_classes = [IsAdmin]
    search_fields = ['usuario__first_name', 'usuario__last_name', 'usuario__username', 'cedula']

    @action(detail=False, methods=['get'])
    @action(detail=False, methods=['get'])
    def reporte_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listado Docentes"

        from openpyxl.styles import Font
        header_font = Font(bold=True)
        title_font = Font(bold=True, size=14)

        ws['A1'] = "LISTADO GENERAL DE DOCENTES"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')

        headers = ['Nombres', 'Apellidos', 'Cédula', 'Teléfono', 'Correo', 'Contratación']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font

        row_num = 4
        for docente in self.filter_queryset(self.get_queryset()):
            ws.cell(row=row_num, column=1, value=docente.usuario.first_name)
            ws.cell(row=row_num, column=2, value=docente.usuario.last_name)
            ws.cell(row=row_num, column=3, value=docente.cedula)
            ws.cell(row=row_num, column=4, value=docente.telefono)
            ws.cell(row=row_num, column=5, value=docente.usuario.email)
            ws.cell(row=row_num, column=6, value=docente.tipo_contratacion)
            row_num += 1

        apply_excel_styling(ws, 3, custom_widths={'A': 20.0, 'B': 20.0})

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_docentes.xlsx'
        wb.save(response)
        return response


class AdminViewSet(viewsets.ModelViewSet):
    from gestion.models import Administrador
    queryset = Administrador.objects.all()
    serializer_class = AdministradorSerializer
    permission_classes = [IsAdmin]
    search_fields = ['usuario__first_name', 'usuario__last_name', 'usuario__username', 'cedula']

    @action(detail=False, methods=['get'])
    @action(detail=False, methods=['get'])
    def reporte_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listado Administradores"

        from openpyxl.styles import Font
        header_font = Font(bold=True)
        title_font = Font(bold=True, size=14)

        ws['A1'] = "LISTADO DE ADMINISTRADORES"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')

        headers = ['Nombres', 'Apellidos', 'Cédula', 'Teléfono', 'Correo']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font

        row_num = 4
        for admin in self.filter_queryset(self.get_queryset()):
            ws.cell(row=row_num, column=1, value=admin.usuario.first_name)
            ws.cell(row=row_num, column=2, value=admin.usuario.last_name)
            ws.cell(row=row_num, column=3, value=admin.cedula)
            ws.cell(row=row_num, column=4, value=admin.telefono)
            ws.cell(row=row_num, column=5, value=admin.usuario.email)
            row_num += 1

        apply_excel_styling(ws, 3)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_administradores.xlsx'
        wb.save(response)
        return response


class PensumViewSet(viewsets.ModelViewSet):
    queryset = Pensum.objects.all()
    serializer_class = PensumSerializer
    parser_classes = [MultiPartParser, FormParser]
    filterset_fields = ['programa']
    search_fields = ['programa__nombre_programa']

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdmin()]
        return []


class PlanificacionViewSet(viewsets.ModelViewSet):
    queryset = Planificacion.objects.all()
    serializer_class = PlanificacionSerializer
    parser_classes = [MultiPartParser, FormParser]
    filterset_fields = ['asignatura', 'asignatura__codigo', 'codigo_especifico']
    search_fields = ['asignatura__nombre_asignatura', 'asignatura__codigo']

    def perform_create(self, serializer):
        asignatura = serializer.validated_data.get('asignatura')
        user = self.request.user
        
        if not user.is_superuser and not user.groups.filter(name='Administrador').exists():
            if not Seccion.objects.filter(asignatura=asignatura, docente=user).exists():
                 from rest_framework.exceptions import PermissionDenied
                 raise PermissionDenied("No estás asignado a esta asignatura.")

        serializer.save(uploaded_by=self.request.user)

    def perform_destroy(self, instance):
        user = self.request.user
        if not user.is_superuser and not user.groups.filter(name='Administrador').exists():
            if not Seccion.objects.filter(asignatura=instance.asignatura, docente=user).exists():
                 from rest_framework.exceptions import PermissionDenied
                 raise PermissionDenied("No estás asignado a esta asignatura.")
        instance.delete()

    def get_permissions(self):
        from rest_framework.permissions import IsAuthenticated
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsDocenteOrAdmin()]
        return []


class DocumentoCalificacionesViewSet(viewsets.ModelViewSet):
    queryset = DocumentoCalificaciones.objects.all()
    serializer_class = DocumentoCalificacionesSerializer
    parser_classes = [MultiPartParser, FormParser]
    filterset_fields = ['estudiante']
    search_fields = ['estudiante__cedula', 'estudiante__usuario__first_name', 'estudiante__usuario__last_name']

    def get_permissions(self):
        if self.action in ['create']:
            return [IsEstudiante()]
        return []

    def perform_create(self, serializer):
        estudiante = serializer.validated_data.get('estudiante')
        if estudiante.usuario != self.request.user and not self.request.user.is_superuser and not self.request.user.groups.filter(name='Docente').exists():
            raise PermissionError('No autorizado')
        doc = serializer.save()
        try:
            doc.estudiante.calcular_avance()
        except Exception:
            pass


class SeccionViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar secciones e inscripción de estudiantes."""
    queryset = Seccion.objects.all()
    serializer_class = SeccionSerializer
    filterset_fields = ['asignatura', 'asignatura__programa', 'docente']

    @action(detail=False, methods=['get'], url_path='master-horario')
    def master_horario(self, request):
        """
        Retorna horarios de todas las secciones, filtrable por programa, semestre y sección.
        Acceso restringido a Docentes y Administradores.
        """
        user = request.user
        
        is_admin = user.is_superuser or user.groups.filter(name='Administrador').exists()
        is_docente = user.groups.filter(name='Docente').exists()
        
        if not is_admin and not is_docente:
            return Response({'error': 'No autorizado.'}, status=status.HTTP_403_FORBIDDEN)
        
        programa_id = request.query_params.get('programa')
        semestre = request.query_params.get('semestre')
        codigo_seccion = request.query_params.get('seccion')
        
        secciones = Seccion.objects.all().select_related(
            'asignatura', 
            'asignatura__programa', 
            'docente'
        ).prefetch_related('horarios')
        
        if programa_id:
            secciones = secciones.filter(asignatura__programa_id=programa_id)
        
        if semestre:
            secciones = secciones.filter(asignatura__semestre=semestre)
            
        if codigo_seccion:
            secciones = secciones.filter(codigo_seccion__icontains=codigo_seccion)
            
        if is_docente and not is_admin:
            secciones = secciones.filter(docente=user)
            
        horario_data = []
        
        for sec in secciones:
            for h in sec.horarios.all():
                horario_data.append({
                    'id': h.id,
                    'dia': h.dia,
                    'hora_inicio': h.hora_inicio.strftime('%H:%M'),
                    'hora_fin': h.hora_fin.strftime('%H:%M'),
                    'asignatura': sec.asignatura.nombre_asignatura,
                    'codigo': sec.asignatura.codigo,
                    'seccion': sec.codigo_seccion,
                    'semestre': sec.asignatura.semestre,
                    'programa': sec.asignatura.programa.nombre_programa,
                    'aula': h.aula,
                    'docente': sec.docente.get_full_name() if sec.docente else 'Sin asignar',
                })
                
        return Response(horario_data)

    @action(detail=False, methods=['get'], url_path='descargar-master-horario')
    def descargar_master_horario(self, request):
        """Genera Excel del Horario Master con filtros (Diseño Exacto Imagen)."""
        user = request.user
        
        is_admin = user.is_superuser or user.groups.filter(name='Administrador').exists()
        is_docente = user.groups.filter(name='Docente').exists()
        
        if not is_admin and not is_docente:
            return Response({'error': 'No autorizado.'}, status=status.HTTP_403_FORBIDDEN)
        
        programa_id = request.query_params.get('programa')
        semestre = request.query_params.get('semestre')
        codigo_seccion = request.query_params.get('seccion')
        
        secciones = Seccion.objects.all().select_related(
            'asignatura', 
            'asignatura__programa', 
            'docente'
        ).prefetch_related('horarios')
        
        program_name = "TODOS LOS PROGRAMAS"
        if programa_id:
            secciones = secciones.filter(asignatura__programa_id=programa_id)
            from gestion.models import Programa
            try:
                prog = Programa.objects.get(pk=programa_id)
                program_name = prog.nombre_programa
            except:
                program_name = "PROGRAMA DESCONOCIDO"
        
        if semestre:
            secciones = secciones.filter(asignatura__semestre=semestre)
        if codigo_seccion:
            secciones = secciones.filter(codigo_seccion__icontains=codigo_seccion)
            
        docente_filter_text = ""
        if is_docente and not is_admin:
            secciones = secciones.filter(docente=user)
            docente_filter_text = f"DOCENTE: {user.get_full_name()}"
            
        grid_data = {day: {} for day in range(1, 7)}
        subjects_list = {}
        
        bloques_data = [
            (1, '07:00 - 07:45'), (2, '07:45 - 08:30'), (3, '08:30 - 09:15'),
            (4, '09:15 - 10:00'), (5, '10:00 - 10:45'), (6, '10:45 - 11:30'),
            (7, '11:30 - 12:15'), (8, '12:15 - 13:00'), (9, '13:00 - 13:45'),
            (10, '13:45 - 14:30'), (11, '14:30 - 15:15'), (12, '15:15 - 16:00'),
            (13, '16:00 - 16:45'), (14, '16:45 - 17:30')
        ]
        
        start_to_block = { label.split(' - ')[0]: bid for bid, label in bloques_data }
        
        for sec in secciones:
            uid = f"{sec.asignatura.codigo}-{sec.codigo_seccion}"
            if uid not in subjects_list:
                subjects_list[uid] = {
                    'codigo': sec.asignatura.codigo,
                    'asignatura': sec.asignatura.nombre_asignatura,
                    'semestre': sec.asignatura.semestre,
                    'seccion': sec.codigo_seccion,
                    'docente': sec.docente.get_full_name() if sec.docente else "SIN ASIGNAR"
                }
            
            for h in sec.horarios.all():
                start_str = h.hora_inicio.strftime('%H:%M')
                
                start_block = start_to_block.get(start_str)
                if not start_block:
                     for bid, label in bloques_data:
                        if label.startswith(start_str):
                            start_block = bid
                            break
                
                if start_block:
                    h_total = (h.hora_fin.hour * 60 + h.hora_fin.minute) - (h.hora_inicio.hour * 60 + h.hora_inicio.minute)
                    num_blocks = max(1, round(h_total / 45))
                    
                    cell_text = f"{sec.asignatura.nombre_asignatura}\n({h.aula or 'N/A'})"
                    
                    for i in range(num_blocks):
                        current_block = start_block + i
                        if current_block > 14: break
                        
                        if current_block in grid_data[h.dia]:
                            existing = grid_data[h.dia][current_block]
                            if existing['uid'] != uid: # Avoid duplicating self if somehow overlap
                                grid_data[h.dia][current_block] = {'text': existing['text'] + "\n-----\n" + cell_text, 'uid': 'COLLISION'}
                        else:
                            grid_data[h.dia][current_block] = {'text': cell_text, 'uid': uid}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Horario Maestro"
        ws.sheet_view.showGridLines = False
        
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        COLOR_HEADER_BG = "4472C4"
        COLOR_HEADER_TXT = "FFFFFF"
        COLOR_CLASS_BG = "CCC0DA"
        COLOR_BORDER = "000000"

        font_title = Font(name='Calibri', size=11, bold=True, color=COLOR_HEADER_TXT)
        font_header_row = Font(name='Calibri', size=11, bold=True, color=COLOR_HEADER_TXT)
        font_cell = Font(name='Calibri', size=10)
        font_cell_bold = Font(name='Calibri', size=10, bold=True)
        
        fill_header = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        fill_class = PatternFill("solid", fgColor=COLOR_CLASS_BG)
        
        border_thin = Side(border_style="thin", color=COLOR_BORDER)
        border_all = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if is_docente and not is_admin:
            ws['A1'] = "HORARIO ACADÉMICO"
            ws['A1'].font = font_title
            ws['A1'].fill = fill_header
            ws['A1'].alignment = align_center
            ws.merge_cells('A1:H1')

            try:
                from gestion.models import PeriodoAcademico
                periodo_obj = PeriodoAcademico.objects.filter(activo=True).first()
                periodo_str = periodo_obj.nombre_periodo if periodo_obj else "Periodo Actual"
                
                docente_profile = user.docente
                cedula = docente_profile.cedula
                tipo_contratacion = docente_profile.tipo_contratacion
            except Exception:
                cedula = "N/A"
                tipo_contratacion = "N/A"
                periodo_str = "Periodo Actual"

            ws['A2'] = f"DOCENTE: {user.get_full_name().upper()}"
            ws['A2'].alignment = align_left
            ws['A2'].border = border_all
            ws.merge_cells('A2:D2')

            ws['E2'] = f"PERÍODO: {periodo_str.upper()}"
            ws['E2'].alignment = align_left
            ws['E2'].border = border_all
            ws.merge_cells('E2:H2')

            ws['A3'] = f"C.I.: {cedula}"
            ws['A3'].alignment = align_left
            ws['A3'].border = border_all
            ws.merge_cells('A3:D3')

            ws['E3'] = f"CONTRATACIÓN: {tipo_contratacion.upper()}"
            ws['E3'].alignment = align_left
            ws['E3'].border = border_all
            ws.merge_cells('E3:H3')

        else:
            ws['A1'] = "HORARIO MAESTRO DE CLASES"
            ws['A1'].font = font_title
            ws['A1'].fill = fill_header
            ws['A1'].alignment = align_center
            ws.merge_cells('A1:H1')

            ws['A2'] = f"Programa: {program_name}"
            ws['A2'].alignment = align_left
            ws['A2'].border = border_all
            ws.merge_cells('A2:D2')
            
            sem_text = f"Semestre: {semestre}" if semestre else "Semestre: TODOS"
            ws['E2'] = sem_text
            ws['E2'].alignment = align_left
            ws['E2'].border = border_all
            ws.merge_cells('E2:H2')

            sec_text = f"Sección: {codigo_seccion}" if codigo_seccion else "Sección: TODAS"
            ws['A3'] = sec_text
            ws['A3'].alignment = align_left
            ws['A3'].border = border_all
            ws.merge_cells('A3:D3')

            doc_text = docente_filter_text if docente_filter_text else "Vista: ADMINISTRADOR"
            ws['E3'] = doc_text
            ws['E3'].alignment = align_left
            ws['E3'].border = border_all
            ws.merge_cells('E3:H3')

        headers = ["BLOQUE", "HORAS", "LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO"]
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=text)
            cell.font = font_header_row
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_all

        start_row = 6
        for i, (bid, time_label) in enumerate(bloques_data):
            row = start_row + i
            c1 = ws.cell(row=row, column=1, value=bid)
            c1.font = font_cell_bold; c1.alignment = align_center; c1.border = border_all
            c2 = ws.cell(row=row, column=2, value=time_label)
            c2.font = font_cell; c2.alignment = align_center; c2.border = border_all

        for d_idx in range(1, 7):
            col_idx = 2 + d_idx
            current_uid = None
            merge_start_row = None
            
            for b_idx in range(1, 15):
                current_row = start_row + (b_idx - 1)
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border_all
                
                data = grid_data[d_idx].get(b_idx)
                
                if data:
                    new_uid = data['uid']
                    text = data['text']
                    
                    if new_uid == current_uid and new_uid != 'COLLISION':
                        pass # Continue merging
                    else:
                        if current_uid is not None:
                            if merge_start_row < current_row - 1:
                                ws.merge_cells(start_row=merge_start_row, start_column=col_idx, end_row=current_row-1, end_column=col_idx)
                        
                        cell.value = text
                        cell.font = font_cell_bold
                        cell.fill = fill_class
                        cell.alignment = align_center
                        current_uid = new_uid
                        merge_start_row = current_row
                        
                    cell.fill = fill_class
                else:
                    if current_uid is not None:
                         if merge_start_row < current_row:
                             ws.merge_cells(start_row=merge_start_row, start_column=col_idx, end_row=current_row-1, end_column=col_idx)
                         current_uid = None
                         merge_start_row = None
            
            if current_uid is not None:
                 if merge_start_row < (start_row + 14):
                     ws.merge_cells(start_row=merge_start_row, start_column=col_idx, end_row=start_row+13, end_column=col_idx)

        det_head_row = 21
        
        ws.column_dimensions['A'].width = 9.50
        ws.column_dimensions['B'].width = 13.50
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        
        headers_def = [
            (1, "N°"), (2, "CÓDIGO"), (3, "ASIGNATURA"), (5, "SEMESTRE"), (6, "SECCIÓN"), (7, "DOCENTE")
        ]
        
        for col, text in headers_def:
            c = ws.cell(row=det_head_row, column=col, value=text)
            c.fill = fill_header; c.font = font_header_row; c.alignment = align_center; c.border = border_all
            
            if col == 3: # Asignatura merge
                ws.merge_cells(start_row=det_head_row, start_column=3, end_row=det_head_row, end_column=4)
                ws.cell(row=det_head_row, column=4).border = border_all
            if col == 7: # Docente merge
                ws.merge_cells(start_row=det_head_row, start_column=7, end_row=det_head_row, end_column=8)
                ws.cell(row=det_head_row, column=8).border = border_all
                
        sorted_subs = sorted(subjects_list.values(), key=lambda x: (x['semestre'], x['codigo'], x['seccion']))
        row_idx = det_head_row + 1
        
        for idx, sub in enumerate(sorted_subs, 1):
            ws.cell(row=row_idx, column=1, value=idx).border = border_all; ws.cell(row=row_idx, column=1).alignment = align_center
            ws.cell(row=row_idx, column=2, value=sub['codigo']).border = border_all; ws.cell(row=row_idx, column=2).alignment = align_center
            ws.cell(row=row_idx, column=3, value=sub['asignatura']).border = border_all; ws.cell(row=row_idx, column=3).alignment = align_center
            ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=4); ws.cell(row=row_idx, column=4).border = border_all
            ws.cell(row=row_idx, column=5, value=sub['semestre']).border = border_all; ws.cell(row=row_idx, column=5).alignment = align_center
            ws.cell(row=row_idx, column=6, value=sub['seccion']).border = border_all; ws.cell(row=row_idx, column=6).alignment = align_center
            ws.cell(row=row_idx, column=7, value=sub['docente']).border = border_all; ws.cell(row=row_idx, column=7).alignment = align_center
            ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=8); ws.cell(row=row_idx, column=8).border = border_all
            
            row_idx += 1
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Horario_Maestro.xlsx"'
        wb.save(response)
        return response
        ws = wb.active
        ws.title = "Horario Maestro"
        
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=16)
        
        border_thin = Side(border_style="thin", color="000000")
        border_all = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        fill_header = PatternFill("solid", fgColor="E0E0E0")
        fill_block = PatternFill("solid", fgColor="F5F5F5")

        ws['A1'] = "REPÚBLICA BOLIVARIANA DE VENEZUELA"
        ws['A2'] = program_name
        ws['A3'] = "HORARIO MAESTRO DE CLASES"
        
        info_parts = []
        if semestre: info_parts.append(f"SEMESTRE: {semestre}")
        if codigo_seccion: info_parts.append(f"SECCIÓN: {codigo_seccion}")
        if docente_filter_text: info_parts.append(docente_filter_text)
        
        ws['A4'] = " | ".join(info_parts) if info_parts else "VISTA GENERAL"
        
        for i in range(1, 5):
            ws.merge_cells(f'A{i}:H{i}')
            ws[f'A{i}'].font = title_font if i == 3 else header_font
            ws[f'A{i}'].alignment = align_center

        headers = ['BLOQUE', 'HORA', 'LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = fill_header
            cell.border = border_all
            cell.alignment = align_center

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        for col in range(3, 9):
            ws.column_dimensions[get_column_letter(col)].width = 25

        bloques_data = [
            (1, '07:00 - 07:45'), (2, '07:45 - 08:30'), (3, '08:30 - 09:15'),
            (4, '09:15 - 10:00'), (5, '10:00 - 10:45'), (6, '10:45 - 11:30'),
            (7, '11:30 - 12:15'), (8, '12:15 - 13:00'), (9, '13:00 - 13:45'),
            (10, '13:45 - 14:30'), (11, '14:30 - 15:15'), (12, '15:15 - 16:00'),
            (13, '16:00 - 16:45'), (14, '16:45 - 17:30')
        ]

        grid_map = {} # (dia, block_id) -> list of strings
        subjects_list = {}

        for sec in secciones:
            uid = f"{sec.asignatura.codigo}-{sec.codigo_seccion}"
            if uid not in subjects_list:
                subjects_list[uid] = {
                    'codigo': sec.asignatura.codigo,
                    'asignatura': sec.asignatura.nombre_asignatura,
                    'semestre': sec.asignatura.semestre,
                    'seccion': sec.codigo_seccion,
                    'docente': sec.docente.get_full_name() if sec.docente else "SIN ASIGNAR"
                }

            for h in sec.horarios.all():
                start_str = h.hora_inicio.strftime('%H:%M')
                block_id = 0
                for bid, label in bloques_data:
                    if label.startswith(start_str):
                        block_id = bid
                        break
                
                if block_id > 0:
                    key = (h.dia, block_id)
                    if key not in grid_map:
                        grid_map[key] = []
                    
                    entry = f"{sec.asignatura.nombre_asignatura}\nSec: {sec.codigo_seccion} | Aula: {h.aula or 'N/A'}"
                    if sec.docente:
                        entry += f"\nProf. {sec.docente.first_name} {sec.docente.last_name}"
                    
                    grid_map[key].append(entry)

        start_row = 7
        for i, (block_id, time_label) in enumerate(bloques_data):
            row_num = start_row + i
            
            ws.cell(row=row_num, column=1, value=f"BLOQUE {block_id}").border = border_all
            ws.cell(row=row_num, column=1).fill = fill_block
            ws.cell(row=row_num, column=1).alignment = align_center

            ws.cell(row=row_num, column=2, value=time_label).border = border_all
            ws.cell(row=row_num, column=2).alignment = align_center
            
            for day_idx in range(1, 7):
                cell_col = 2 + day_idx
                entries = grid_map.get((day_idx, block_id), [])
                
                cell_content = "\n\n".join(entries)
                cell = ws.cell(row=row_num, column=cell_col, value=cell_content)
                cell.border = border_all
                cell.alignment = align_center
                cell.font = Font(size=9)
                if entries:
                    cell.fill = PatternFill("solid", fgColor="E6E6FA")

        detail_start_row = start_row + 15
        
        ws[f'A{detail_start_row}'] = "DETALLE DE SECCIONES ENCONTRADAS"
        ws[f'A{detail_start_row}'].font = header_font
        ws.merge_cells(f'A{detail_start_row}:F{detail_start_row}')
        
        detail_header_row = detail_start_row + 1
        d_headers = ['N°', 'CÓDIGO', 'ASIGNATURA', 'SEM', 'SECCIÓN', 'DOCENTE']
        d_widths = [5, 15, 35, 5, 10, 30]
        
        for col, (header, width) in enumerate(zip(d_headers, d_widths), 1):
            cell = ws.cell(row=detail_header_row, column=col, value=header)
            cell.font = header_font
            cell.border = border_all
            cell.fill = fill_header
            
        r_idx = detail_header_row + 1
        sorted_subs = sorted(subjects_list.values(), key=lambda x: (x['semestre'], x['codigo'], x['seccion']))
        
        for idx, sub in enumerate(sorted_subs, 1):
            ws.cell(row=r_idx, column=1, value=idx).border = border_all
            ws.cell(row=r_idx, column=2, value=sub['codigo']).border = border_all
            ws.cell(row=r_idx, column=3, value=sub['asignatura']).border = border_all
            ws.cell(row=r_idx, column=4, value=sub['semestre']).border = border_all
            ws.cell(row=r_idx, column=5, value=sub['seccion']).border = border_all
            ws.cell(row=r_idx, column=6, value=sub['docente']).border = border_all
            
            ws.cell(row=r_idx, column=1).alignment = align_center
            ws.cell(row=r_idx, column=2).alignment = align_center
            ws.cell(row=r_idx, column=4).alignment = align_center
            ws.cell(row=r_idx, column=5).alignment = align_center
            r_idx += 1
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Horario_Maestro.xlsx"'
        wb.save(response)
        return response
        ws = wb.active
        ws.title = "Horario Maestro"
        
        from openpyxl.styles import Font, Alignment
        header_font = Font(bold=True)
        
        headers = ['Día', 'Hora Inicio', 'Hora Fin', 'Código', 'Asignatura', 'Sección', 'Semestre', 'Programa', 'Aula', 'Docente']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            
        row_num = 2
        for sec in secciones:
            for h in sec.horarios.all():
                time_map = {1: '07:00', 2: '07:45', 3: '08:30', 4: '09:15', 5: '10:00', 6: '10:45', 7: '11:30', 8: '12:15', 9: '13:00', 10: '13:45', 11: '14:30', 12: '15:15', 13: '16:00', 14: '16:45'}
                
                day_map = {1: 'LUNES', 2: 'MARTES', 3: 'MIÉRCOLES', 4: 'JUEVES', 5: 'VIERNES', 6: 'SÁBADO'}
                dia_str = day_map.get(h.dia, str(h.dia))
                
                ws.cell(row=row_num, column=1, value=dia_str)
                ws.cell(row=row_num, column=2, value=h.hora_inicio.strftime('%H:%M'))
                ws.cell(row=row_num, column=3, value=h.hora_fin.strftime('%H:%M'))
                ws.cell(row=row_num, column=4, value=sec.asignatura.codigo)
                ws.cell(row=row_num, column=5, value=sec.asignatura.nombre_asignatura)
                ws.cell(row=row_num, column=6, value=sec.codigo_seccion)
                ws.cell(row=row_num, column=7, value=sec.asignatura.semestre)
                ws.cell(row=row_num, column=8, value=sec.asignatura.programa.nombre_programa)
                ws.cell(row=row_num, column=9, value=h.aula or 'Sin Aula')
                ws.cell(row=row_num, column=10, value=sec.docente.get_full_name() if sec.docente else 'Sin Asignar')
                row_num += 1
                
        dims = {'A': 15, 'B': 10, 'C': 10, 'D': 12, 'E': 40, 'F': 10, 'G': 10, 'H': 30, 'I': 15, 'J': 25}
        for col, width in dims.items():
            ws.column_dimensions[col].width = width
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Horario_Maestro.xlsx"'
        wb.save(response)
        return response

    def get_permissions(self):
        from rest_framework.permissions import IsAuthenticated
        if self.action in ['list', 'retrieve', 'estudiantes']:
            return [IsAuthenticated()]
        if self.action in ['inscribir_estudiante', 'desinscribir_estudiante', 'descargar_listado', 'master_horario', 'descargar_master_horario']:
            return [IsDocenteOrAdmin()]
        if self.action in ['inscribirme', 'desinscribirme']:
            return [IsEstudiante()]
        if self.action in ['mis_secciones', 'calificar']:
            return [IsDocente()]
        return [IsAdmin()]

    @action(detail=True, methods=['get'], url_path='estudiantes')
    def estudiantes(self, request, pk=None):
        """Lista los estudiantes inscritos en esta sección."""
        seccion = self.get_object()
        from gestion.api.serializers import DetalleInscripcionListSerializer
        detalles = seccion.estudiantes_inscritos.filter(estatus='CURSANDO')
        serializer = DetalleInscripcionListSerializer(detalles, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='descargar-listado')
    def descargar_listado(self, request, pk=None):
        """Descarga el listado de estudiantes inscritos en la sección como Excel."""
        seccion = self.get_object()
        user = request.user
        
        if not user.is_superuser and not user.groups.filter(name='Administrador').exists():
            if not user.groups.filter(name='Docente').exists():
                return Response({'error': 'No autorizado.'}, status=status.HTTP_403_FORBIDDEN)
            if seccion.docente != user:
                return Response({'error': 'No estás asignado a esta sección.'}, status=status.HTTP_403_FORBIDDEN)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Listado {seccion.codigo_seccion}"

        from openpyxl.styles import Font, Alignment
        header_font = Font(bold=True)
        title_font = Font(bold=True, size=14)
        
        ws['A1'] = "PLANILLA DE EVALUACIÓN"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = "Asignatura:"
        ws['B2'] = f"{seccion.asignatura.nombre_asignatura} ({seccion.asignatura.codigo})"
        ws['A2'].font = header_font
        ws.merge_cells('B2:E2')
        
        ws['A3'] = "Sección:"
        ws['B3'] = seccion.codigo_seccion
        ws['A3'].font = header_font
        
        ws['A4'] = "Docente:"
        ws['B4'] = seccion.docente.get_full_name() if seccion.docente else "Sin Asignar"
        ws['A4'].font = header_font
        ws.merge_cells('B4:E4')

        from gestion.models import PeriodoAcademico
        periodo = PeriodoAcademico.objects.filter(activo=True).first()
        ws['A5'] = "Período:"
        ws['B5'] = str(periodo) if periodo else "N/A"
        ws['A5'].font = header_font
        ws.merge_cells('B5:E5')
        
        headers = ['Cédula', 'Nombre', 'Apellido', 'Correo', 'Nota 1', 'Nota 2', 'Nota 3', 'Nota 4', 'Nota R', 'Nota Final', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
        
        row_num = 8
        detalles = seccion.estudiantes_inscritos.select_related('inscripcion__estudiante__usuario').order_by('inscripcion__estudiante__usuario__last_name')
        
        for detalle in detalles:
            est = detalle.inscripcion.estudiante
            ws.cell(row=row_num, column=1, value=est.cedula)
            ws.cell(row=row_num, column=2, value=est.usuario.first_name)
            ws.cell(row=row_num, column=3, value=est.usuario.last_name)
            ws.cell(row=row_num, column=4, value=est.usuario.email)
            ws.cell(row=row_num, column=5, value=float(detalle.nota1) if detalle.nota1 is not None else '')
            ws.cell(row=row_num, column=6, value=float(detalle.nota2) if detalle.nota2 is not None else '')
            ws.cell(row=row_num, column=7, value=float(detalle.nota3) if detalle.nota3 is not None else '')
            ws.cell(row=row_num, column=8, value=float(detalle.nota4) if detalle.nota4 is not None else '')
            ws.cell(row=row_num, column=9, value=float(detalle.nota_reparacion) if detalle.nota_reparacion is not None else '')
            ws.cell(row=row_num, column=10, value=float(detalle.nota_final) if detalle.nota_final is not None else '')
            ws.cell(row=row_num, column=11, value=detalle.estatus)
            row_num += 1
            
        apply_excel_styling(ws, 7)
        
        filename = f"listado_{seccion.asignatura.codigo}_{seccion.codigo_seccion}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    @action(detail=True, methods=['post'], url_path='inscribir-estudiante')
    def inscribir_estudiante(self, request, pk=None):
        """Permite a un docente inscribir un estudiante en su sección asignada."""
        seccion = self.get_object()
        user = request.user
        
        if not user.is_superuser and not user.groups.filter(name='Administrador').exists():
            if seccion.docente != user:
                return Response(
                    {'error': 'No estás asignado a esta sección.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        estudiante_id = request.data.get('estudiante_id')
        if not estudiante_id:
            return Response({'error': 'Se requiere estudiante_id.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            estudiante = Estudiante.objects.get(pk=estudiante_id)
        except Estudiante.DoesNotExist:
            return Response({'error': 'Estudiante no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        if estudiante.programa_id != seccion.asignatura.programa_id:
            return Response(
                {'error': 'El estudiante no pertenece al programa de esta asignatura.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        result = self._inscribir_estudiante_en_seccion(estudiante, seccion)
        if result.get('error'):
            return Response({'error': result['error']}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({'status': 'Estudiante inscrito exitosamente.', 'detalle_id': result.get('detalle_id')})

    @action(detail=True, methods=['post'], url_path='inscribirme')
    def inscribirme(self, request, pk=None):
        """Permite a un estudiante auto-inscribirse en una sección de su programa."""
        seccion = self.get_object()
        user = request.user

        try:
            estudiante = Estudiante.objects.get(usuario=user)
        except Estudiante.DoesNotExist:
            return Response({'error': 'No tienes perfil de estudiante.'}, status=status.HTTP_400_BAD_REQUEST)

        if estudiante.programa_id != seccion.asignatura.programa_id:
            return Response(
                {'error': 'Esta asignatura no pertenece a tu programa.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        result = self._inscribir_estudiante_en_seccion(estudiante, seccion)
        if result.get('error'):
            return Response({'error': result['error']}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({'status': 'Te has inscrito exitosamente.', 'detalle_id': result.get('detalle_id')})

    @action(detail=True, methods=['post'], url_path='desinscribirme')
    def desinscribirme(self, request, pk=None):
        """Permite a un estudiante desinscribirse de una sección."""
        seccion = self.get_object()
        user = request.user

        try:
            estudiante = Estudiante.objects.get(usuario=user)
        except Estudiante.DoesNotExist:
            return Response({'error': 'No tienes perfil de estudiante.'}, status=status.HTTP_400_BAD_REQUEST)

        from gestion.models import Inscripcion, DetalleInscripcion, PeriodoAcademico
        
        detalle = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante,
            seccion=seccion,
            estatus='CURSANDO'
        ).first()

        if not detalle:
            return Response({'error': 'No estás inscrito en esta sección.'}, status=status.HTTP_400_BAD_REQUEST)

        detalle.delete()
        return Response({'status': 'Te has desinscrito exitosamente.'})

    @action(detail=True, methods=['post'], url_path='desinscribir-estudiante')
    def desinscribir_estudiante(self, request, pk=None):
        """Permite a un docente/admin desinscribir un estudiante de su sección."""
        seccion = self.get_object()
        user = request.user

        if not user.is_superuser and not user.groups.filter(name='Administrador').exists():
            if seccion.docente != user:
                return Response(
                    {'error': 'No estás asignado a esta sección.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        estudiante_id = request.data.get('estudiante_id')
        if not estudiante_id:
            return Response({'error': 'Se requiere estudiante_id.'}, status=status.HTTP_400_BAD_REQUEST)

        from gestion.models import DetalleInscripcion
        detalle = DetalleInscripcion.objects.filter(
            inscripcion__estudiante_id=estudiante_id,
            seccion=seccion
        ).first()

        if not detalle:
            return Response({'error': 'Estudiante no encontrado en esta sección.'}, status=status.HTTP_404_NOT_FOUND)

        detalle.delete()
        return Response({'status': 'Estudiante desinscrito exitosamente.'})

    @action(detail=False, methods=['get'], url_path='mis-secciones')
    def mis_secciones(self, request):
        """Devuelve las secciones asignadas al docente autenticado con sus estudiantes."""
        user = request.user
        
        if not user.groups.filter(name='Docente').exists() and not user.is_superuser and not user.groups.filter(name='Administrador').exists():
            return Response({'error': 'Solo los docentes y administradores pueden acceder a este recurso.'}, status=status.HTTP_403_FORBIDDEN)
        
        SPECIAL_CODES = ['TAI-01', 'PRO-01', 'PSI-30010']
        if user.is_superuser or user.groups.filter(name='Administrador').exists():
            from gestion.models import Asignatura
            for code in SPECIAL_CODES:
                asig = Asignatura.objects.filter(codigo=code).first()
                if asig:
                    if not Seccion.objects.filter(asignatura=asig).exists():
                         Seccion.objects.create(asignatura=asig, codigo_seccion='01', docente=None)

            secciones = Seccion.objects.all().select_related('asignatura', 'asignatura__programa', 'docente')
        else:
            secciones = Seccion.objects.filter(docente=user).exclude(asignatura__codigo__in=SPECIAL_CODES).select_related('asignatura', 'asignatura__programa', 'docente')

        programa_id = request.query_params.get('programa')
        if programa_id:
            secciones = secciones.filter(asignatura__programa_id=programa_id)
        
        result = []
        for seccion in secciones:
            estudiantes_data = []
            detalles = DetalleInscripcion.objects.filter(seccion=seccion).select_related(
                'inscripcion__estudiante', 'inscripcion__estudiante__usuario'
            )
            for detalle in detalles:
                est = detalle.inscripcion.estudiante
                estudiantes_data.append({
                    'detalle_id': detalle.id,
                    'estudiante_id': est.id,
                    'cedula': est.cedula,
                    'nombre': est.usuario.get_full_name(),
                    'nota1': float(detalle.nota1) if detalle.nota1 else None,
                    'nota2': float(detalle.nota2) if detalle.nota2 else None,
                    'nota3': float(detalle.nota3) if detalle.nota3 else None,
                    'nota4': float(detalle.nota4) if detalle.nota4 else None,
                    'nota_reparacion': float(detalle.nota_reparacion) if detalle.nota_reparacion else None,
                    'nota_final': float(detalle.nota_final) if detalle.nota_final else None,
                    'estatus': detalle.estatus
                })
            
            result.append({
                'id': seccion.id,
                'codigo_seccion': seccion.codigo_seccion,
                'asignatura_id': seccion.asignatura.id,
                'asignatura_codigo': seccion.asignatura.codigo,
                'asignatura_nombre': seccion.asignatura.nombre_asignatura,
                'programa': seccion.asignatura.programa.nombre_programa,
                'semestre': seccion.asignatura.semestre,
                'docente_nombre': seccion.docente.get_full_name() if seccion.docente else 'Sin Docente',
                'estudiantes': estudiantes_data,
                'total_estudiantes': len(estudiantes_data)
            })
        
        return Response(result)

    @action(detail=True, methods=['post'], url_path='calificar')
    def calificar(self, request, pk=None):
        """Permite a un docente asignar calificaciones a un estudiante."""
        seccion = self.get_object()
        user = request.user
        
        if not user.is_superuser and not user.groups.filter(name='Administrador').exists():
            if seccion.docente != user:
                return Response({'error': 'No estás asignado a esta sección.'}, status=status.HTTP_403_FORBIDDEN)
        
        detalle_id = request.data.get('detalle_id')
        nota1 = request.data.get('nota1')
        nota2 = request.data.get('nota2')
        nota3 = request.data.get('nota3')
        nota4 = request.data.get('nota4')
        nota_reparacion = request.data.get('nota_reparacion')
        
        if not detalle_id:
            return Response({'error': 'Se requiere detalle_id.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            detalle = DetalleInscripcion.objects.get(pk=detalle_id, seccion=seccion)
        except DetalleInscripcion.DoesNotExist:
            return Response({'error': 'Detalle de inscripción no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        
        from decimal import Decimal, InvalidOperation
        def parse_nota(val):
            if val is None or val == '':
                return None
            try:
                nota = Decimal(str(val))
                if nota < 1 or nota > 20:
                    raise ValueError('Nota fuera de rango')
                return nota
            except (InvalidOperation, ValueError):
                return None
        
        if nota1 is not None:
            detalle.nota1 = parse_nota(nota1)
        if nota2 is not None:
            detalle.nota2 = parse_nota(nota2)
        if nota3 is not None:
            detalle.nota3 = parse_nota(nota3)
        if nota4 is not None:
            detalle.nota4 = parse_nota(nota4)
        
        if nota_reparacion is not None:
            if nota_reparacion == '' or nota_reparacion is None:
                detalle.nota_reparacion = None
            else:
                notas_actuales = [
                    detalle.nota1 if nota1 is None else parse_nota(nota1),
                    detalle.nota2 if nota2 is None else parse_nota(nota2),
                    detalle.nota3 if nota3 is None else parse_nota(nota3),
                    detalle.nota4 if nota4 is None else parse_nota(nota4)
                ]
                
                if all(n is not None for n in notas_actuales):
                    from decimal import Decimal
                    promedio_temp = sum(notas_actuales) / Decimal(4)
                    
                    if promedio_temp >= 10:
                        return Response({
                            'error': 'Solo se puede cargar Nota de Reparación para estudiantes reprobados (promedio < 10).'
                        }, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({
                        'error': 'Debe cargar las 4 notas parciales antes de asignar una Nota de Reparación.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                detalle.nota_reparacion = parse_nota(nota_reparacion)
        
        detalle.save()  # El método save() calcula nota_final automáticamente
        
        return Response({
            'status': 'Calificaciones guardadas.',
            'nota_final': float(detalle.nota_final) if detalle.nota_final else None,
            'nota_reparacion': float(detalle.nota_reparacion) if detalle.nota_reparacion else None,
            'estatus': detalle.estatus
        })

        return {'detalle_id': detalle.id}

    def _inscribir_estudiante_en_seccion(self, estudiante, seccion, allow_conflicts=False):
        """Método interno para realizar la inscripción con validaciones."""
        from gestion.models import Inscripcion, DetalleInscripcion, PeriodoAcademico, Asignatura, Horario
        from django.core.exceptions import ValidationError
        from django.db.models import Sum, Q

        asignatura = seccion.asignatura
        user_requesting = self.request.user

        
        should_check_conflicts = True
        if user_requesting.is_superuser or user_requesting.groups.filter(name='Administrador').exists() or user_requesting.groups.filter(name='Docente').exists():
            should_check_conflicts = False

        nombre_upper = asignatura.nombre_asignatura.upper()
        if "SERVICIO COMUNITARIO" in nombre_upper:
            total_creditos = Asignatura.objects.filter(programa=estudiante.programa).aggregate(total=Sum('creditos'))['total'] or 0
            
            creditos_aprobados = DetalleInscripcion.objects.filter(
                inscripcion__estudiante=estudiante
            ).filter(
                Q(estatus='APROBADO') | Q(nota_final__gte=10)
            ).aggregate(total=Sum('asignatura__creditos'))['total'] or 0

            if creditos_aprobados < (total_creditos * 0.5):
                porcentaje_actual = (creditos_aprobados / total_creditos * 100) if total_creditos > 0 else 0
                return {
                    'error': f'Requisito no cumplido: Para cursar Servicio Comunitario debes tener aprobado el 50% de las Unidades de Crédito. (Tienes: {creditos_aprobados} UC - {porcentaje_actual:.1f}%)'
                }

        periodo = PeriodoAcademico.objects.filter(activo=True, inscripciones_activas=True).first()
        if not periodo:
            periodo_cerrado = PeriodoAcademico.objects.filter(activo=True).first()
            if periodo_cerrado:
                return {'error': f'Las inscripciones están cerradas para el período {periodo_cerrado.nombre_periodo}.'}
            return {'error': 'No hay período académico activo.'}

        ya_inscrito_periodo = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante,
            inscripcion__periodo=periodo,
            asignatura=asignatura
        ).exists()

        if ya_inscrito_periodo:
            return {'error': 'El estudiante ya está inscrito en esta asignatura en el período actual.'}

        ya_aprobo = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante,
            asignatura=asignatura,
            estatus='APROBADO'
        ).exists()

        if ya_aprobo:
            return {'error': 'El estudiante ya aprobó esta asignatura anteriormente.'}

        prelaciones = asignatura.prelaciones.all()
        for prereq in prelaciones:
            from django.db.models import Q
            aprobada = DetalleInscripcion.objects.filter(
                inscripcion__estudiante=estudiante,
                asignatura=prereq
            ).filter(Q(nota_final__gte=10) | Q(estatus='APROBADO')).exists()
            
            if not aprobada:
                return {'error': f'No puedes inscribir esta materia. Requiere haber aprobado: {prereq.nombre_asignatura} ({prereq.codigo}).'}

        uc_inscritas = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante,
            inscripcion__periodo=periodo
        ).aggregate(total=Sum('asignatura__creditos'))['total'] or 0

        uc_nueva = asignatura.creditos
        
        if (uc_inscritas + uc_nueva) > 35:
            return {
                'error': f'No puedes inscribir esta asignatura porque superarías el límite de 35 UC. (Tienes {uc_inscritas} UC + {uc_nueva} UC de esta materia = {uc_inscritas + uc_nueva})'
            }

        if should_check_conflicts:
            horarios_nuevos = seccion.horarios.all() # QuerySet de Horario
            if horarios_nuevos.exists():
                inscripciones_actuales = DetalleInscripcion.objects.filter(
                    inscripcion__estudiante=estudiante,
                    inscripcion__periodo=periodo
                ).select_related('seccion', 'asignatura')

                for inscripcion_existente in inscripciones_actuales:
                    seccion_existente = inscripcion_existente.seccion
                    if not seccion_existente: continue
                    
                    horarios_existentes = seccion_existente.horarios.all()

                    for h_nuevo in horarios_nuevos:
                        for h_existente in horarios_existentes:
                            if h_nuevo.dia == h_existente.dia:
                                if (h_nuevo.hora_inicio < h_existente.hora_fin) and (h_nuevo.hora_fin > h_existente.hora_inicio):
                                    dia_str = h_nuevo.get_dia_display()
                                    return {
                                        'error': (
                                            f"CHOQUE DE HORARIO: La asignatura '{asignatura.nombre_asignatura}' "
                                            f"choca con '{seccion_existente.asignatura.nombre_asignatura}' "
                                            f"el día {dia_str} ({h_nuevo.hora_inicio.strftime('%H:%M')} - {h_nuevo.hora_fin.strftime('%H:%M')}). "
                                            f"Por favor comunica esta situación a un administrador."
                                        )
                                    }

        inscripcion, _ = Inscripcion.objects.get_or_create(
            estudiante=estudiante,
            periodo=periodo
        )

        detalle = DetalleInscripcion.objects.create(
            inscripcion=inscripcion,
            asignatura=asignatura,
            seccion=seccion,
            estatus='CURSANDO'
        )

        return {'detalle_id': detalle.id}


class PeriodoAcademicoViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar períodos académicos."""
    queryset = PeriodoAcademico.objects.all().order_by('-anio', '-nombre_periodo')
    serializer_class = PeriodoAcademicoSerializer

    def get_permissions(self):
        from rest_framework.permissions import IsAuthenticated
        if self.action in ['list', 'retrieve', 'activo']:
            return [IsAuthenticated()]
        return [IsAdmin()]

    @action(detail=False, methods=['get'], url_path='activo')
    def activo(self, request):
        """Retorna el período académico activo actual."""
        periodo = PeriodoAcademico.objects.filter(activo=True).first()
        if periodo:
            return Response(PeriodoAcademicoSerializer(periodo).data)
        return Response({'error': 'No hay período activo.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='toggle-inscripciones')
    def toggle_inscripciones(self, request, pk=None):
        """Activa o desactiva las inscripciones para un período."""
        from datetime import date
        periodo = self.get_object()
        hoy = date.today()
        
        if periodo.fecha_fin < hoy:
            return Response({
                'error': 'No se pueden modificar inscripciones de un período que ya finalizó.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        periodo.inscripciones_activas = not periodo.inscripciones_activas
        periodo.save()
        return Response({
            'status': 'success',
            'inscripciones_activas': periodo.inscripciones_activas,
            'mensaje': f'Inscripciones {"activadas" if periodo.inscripciones_activas else "desactivadas"} para {periodo.nombre_periodo}'
        })

    @action(detail=True, methods=['post'], url_path='activar')
    def activar_periodo(self, request, pk=None):
        """Establece este período como el activo (desactiva los demás)."""
        from datetime import date
        periodo = self.get_object()
        hoy = date.today()
        
        if periodo.fecha_fin < hoy:
            return Response({
                'error': 'No se puede activar un período que ya finalizó.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if periodo.fecha_inicio > hoy:
            return Response({
                'error': f'Este período no puede activarse hasta {periodo.fecha_inicio}.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        PeriodoAcademico.objects.update(activo=False)
        periodo.activo = True
        periodo.save()
        return Response({
            'status': 'success',
            'mensaje': f'Período {periodo.nombre_periodo} activado correctamente.'
        })


class EstadisticasViewSet(viewsets.ViewSet):
    """ViewSet para estadísticas académicas."""
    permission_classes = [IsDocenteOrAdmin]
    
    def get_permissions(self):
        if self.action == 'mi_progreso':
            return [IsEstudiante()]
        return [IsDocenteOrAdmin()]

    @action(detail=False, methods=['get'], url_path='descargar-desglose-excel')
    def descargar_desglose_excel(self, request):
        """Genera reporte Excel del desglose académico (Admin/Docente)."""
        from django.db.models import Avg, Max, Min, Count
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from gestion.models import Programa
        
        programa_id = request.query_params.get('programa')
        user = request.user
        
        is_admin = user.is_superuser or user.groups.filter(name='Administrador').exists()
        is_docente = user.groups.filter(name='Docente').exists()
        
        if not is_admin and not is_docente:
            return Response({'error': 'No autorizado.'}, status=status.HTTP_403_FORBIDDEN)

        asignaturas = Asignatura.objects.all().order_by('semestre', 'orden')
        if programa_id:
            asignaturas = asignaturas.filter(programa_id=programa_id)
            try:
                programa_nombre = Programa.objects.get(id=programa_id).nombre_programa
            except:
                programa_nombre = "Desconocido"
        else:
            programa_nombre = "Todos"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Desglose Académico"
        
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        
        ws['A1'] = "MONITOREO DE AVANCE EDUCATIVO"
        ws['A1'].font = title_font
        ws.merge_cells('A1:I1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = "Carrera:"
        ws['B2'] = programa_nombre
        ws['A2'].font = header_font
        ws.merge_cells('B2:E2')
        
        periodo_actual = PeriodoAcademico.objects.filter(activo=True).first()
        ws['A3'] = "Período:"
        ws['B3'] = str(periodo_actual) if periodo_actual else "N/A"
        ws['A3'].font = header_font
        
        headers = ['Semestre', 'Código', 'Asignatura', 'Sección', 'Docente', 'Inscritos', 'Promedio', 'Máxima', 'Mínima', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            
        row_num = 6
        
        for asig in asignaturas:
            secciones = asig.secciones.filter(docente__isnull=False).select_related('docente')
            
            if is_docente and not is_admin:
                secciones = secciones.filter(docente=user)
                
            for seccion in secciones:
                stats = DetalleInscripcion.objects.filter(seccion=seccion).aggregate(
                    count=Count('id'),
                    avg=Avg('nota_final'),
                    max=Max('nota_final'),
                    min=Min('nota_final')
                )
                
                
                count = stats['count'] or 0
                avg = round(float(stats['avg']), 2) if stats['avg'] else 0
                max_val = round(float(stats['max']), 2) if stats['max'] else 0
                min_val = round(float(stats['min']), 2) if stats['min'] else 0
                
                ws.cell(row=row_num, column=1, value=asig.semestre)
                ws.cell(row=row_num, column=2, value=asig.codigo)
                ws.cell(row=row_num, column=3, value=asig.nombre_asignatura)
                ws.cell(row=row_num, column=4, value=seccion.codigo_seccion)
                ws.cell(row=row_num, column=5, value=seccion.docente.get_full_name() if seccion.docente else 'Sin asignar')
                ws.cell(row=row_num, column=6, value=count)
                ws.cell(row=row_num, column=7, value=avg)
                ws.cell(row=row_num, column=8, value=max_val)
                ws.cell(row=row_num, column=9, value=min_val)
                ws.cell(row=row_num, column=10, value="Activo" if count > 0 else "Sin Estudiantes") # Estado derivado simple
                row_num += 1

        apply_excel_styling(ws, 5, custom_widths={'A': 10.0, 'B': 14.0, 'C': 42.0, 'E': 42.0})
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Desglose_Academico.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='desglose')
    def desglose(self, request):
        """Devuelve el desglose académico real por semestre/asignatura/sección."""
        from django.db.models import Avg, Max, Min, Count
        
        programa_id = request.query_params.get('programa')
        user = request.user
        
        is_admin = user.is_superuser or user.groups.filter(name='Administrador').exists()
        is_docente = user.groups.filter(name='Docente').exists()
        
        asignaturas = Asignatura.objects.all().order_by('semestre', 'orden')
        if programa_id:
            asignaturas = asignaturas.filter(programa_id=programa_id)
        
        semestres = {}
        for asig in asignaturas:
            sem_key = asig.semestre
            if sem_key not in semestres:
                semestres[sem_key] = {
                    'id': sem_key,
                    'name': f'Semestre {sem_key}',
                    'subjects': []
                }
            
            secciones_data = []
            secciones = asig.secciones.filter(docente__isnull=False).select_related('docente')
            
            if is_docente and not is_admin:
                secciones = secciones.filter(docente=user)
            
            for seccion in secciones:
                stats = DetalleInscripcion.objects.filter(
                    seccion=seccion
                ).aggregate(
                    count=Count('id'),
                    avg=Avg('nota_final'),
                    max=Max('nota_final'),
                    min=Min('nota_final')
                )
                
                secciones_data.append({
                    'id': seccion.id,
                    'code': seccion.codigo_seccion,
                    'docente': seccion.docente.get_full_name() if seccion.docente else 'Sin asignar',
                    'count': stats['count'] or 0,
                    'avg': round(float(stats['avg']), 2) if stats['avg'] else 0,
                    'max': round(float(stats['max']), 2) if stats['max'] else 0,
                    'min': round(float(stats['min']), 2) if stats['min'] else 0
                })
            
            if secciones_data:  # Solo agregar asignaturas con secciones activas
                semestres[sem_key]['subjects'].append({
                    'id': asig.id,
                    'code': asig.codigo,
                    'name': asig.nombre_asignatura,
                    'sections': secciones_data
                })
        
        result = [sem for sem in sorted(semestres.values(), key=lambda x: x['id']) if sem['subjects']]
        return Response(result)

    @action(detail=False, methods=['get'], url_path='chart-data')
    def chart_data(self, request):
        """Devuelve datos para el gráfico radar según el rol del usuario."""
        from django.db.models import Avg
        
        user = request.user
        is_admin = user.is_superuser or user.groups.filter(name='Administrador').exists()
        is_docente = user.groups.filter(name='Docente').exists()
        programa_id = request.query_params.get('programa')
        
        labels = []
        data = []
        
        for sem_num in range(1, 9):
            labels.append(f'Sem {sem_num}')
            
            if is_admin:
                detalles = DetalleInscripcion.objects.filter(
                    asignatura__semestre=sem_num,
                    nota_final__isnull=False
                )
                if programa_id:
                    detalles = detalles.filter(asignatura__programa_id=programa_id)
                
                avg = detalles.aggregate(avg=Avg('nota_final'))['avg']
                data.append(round(float(avg), 2) if avg else 0)
            
            elif is_docente:
                secciones_docente = Seccion.objects.filter(docente=user, asignatura__semestre=sem_num)
                avg = DetalleInscripcion.objects.filter(
                    seccion__in=secciones_docente,
                    nota_final__isnull=False
                ).aggregate(avg=Avg('nota_final'))['avg']
                
                data.append(round(float(avg), 2) if avg else 0)
            else:
                data.append(0)
        
        return Response({
            'labels': labels,
            'data': data
        })

    @action(detail=False, methods=['get'], url_path='mi-progreso')
    def mi_progreso(self, request):
        """Devuelve el progreso académico del estudiante autenticado."""
        from django.db.models import Avg
        
        user = request.user
        
        try:
            estudiante = Estudiante.objects.get(usuario=user)
        except Estudiante.DoesNotExist:
            return Response({'error': 'Usuario no es estudiante.'}, status=status.HTTP_400_BAD_REQUEST)
        
        avance = estudiante.calcular_avance()
        total_asignaturas = estudiante.programa.asignatura_set.count() if estudiante.programa else 0
        aprobadas = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante,
            nota_final__gte=10
        ).count()
        
        nombre = ''
        try:
            nombre = estudiante.usuario.get_full_name()
        except Exception:
            nombre = str(estudiante.usuario)
        
        labels = []
        chart_data = []
        
        desglose = {}
        
        detalles = DetalleInscripcion.objects.filter(
            inscripcion__estudiante=estudiante
        ).select_related('asignatura', 'seccion')
        
        for detalle in detalles:
            sem_key = detalle.asignatura.semestre
            if sem_key not in desglose:
                desglose[sem_key] = {
                    'id': sem_key,
                    'name': f'Semestre {sem_key}',
                    'subjects': []
                }
            
            existing_subject = None
            for subj in desglose[sem_key]['subjects']:
                if subj['id'] == detalle.asignatura.id:
                    existing_subject = subj
                    break
            
            section_data = {
                'id': detalle.seccion.id if detalle.seccion else detalle.id,
                'code': detalle.seccion.codigo_seccion if detalle.seccion else 'N/A',
                'nota1': float(detalle.nota1) if detalle.nota1 else None,
                'nota2': float(detalle.nota2) if detalle.nota2 else None,
                'nota3': float(detalle.nota3) if detalle.nota3 else None,
                'nota4': float(detalle.nota4) if detalle.nota4 else None,
                'nota_final': float(detalle.nota_final) if detalle.nota_final else None,
                'estatus': detalle.estatus
            }
            
            if existing_subject:
                existing_subject['sections'].append(section_data)
            else:
                desglose[sem_key]['subjects'].append({
                    'id': detalle.asignatura.id,
                    'code': detalle.asignatura.codigo,
                    'name': detalle.asignatura.nombre_asignatura,
                    'sections': [section_data]
                })
        
        for sem_num in range(1, 9):
            labels.append(f'Sem {sem_num}')
            if sem_num in desglose:
                notas_sem = []
                for subj in desglose[sem_num]['subjects']:
                    for sec in subj['sections']:
                        if sec['nota_final']:
                            notas_sem.append(sec['nota_final'])
                promedio = sum(notas_sem) / len(notas_sem) if notas_sem else 0
                chart_data.append(round(promedio, 2))
            else:
                chart_data.append(0)
        
        return Response({
            'labels': labels,
            'data': chart_data,
            'desglose': sorted(desglose.values(), key=lambda x: x['id']),
            'estudiante': {
                'id': estudiante.id,
                'nombre': nombre,
                'programa': estudiante.programa.nombre_programa if estudiante.programa else '',
                'porcentaje_avance': avance,
                'total_asignaturas': total_asignaturas,
                'aprobadas': aprobadas
            }
        })

