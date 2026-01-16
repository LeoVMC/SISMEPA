import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def apply_excel_styling(ws, header_row_num, custom_widths=None):
    """
    Applies standard professional styling to an Excel worksheet.
    - Title (Row 1): Styled Box (Gray Fill, Bold, Border), Auto-merged to table width.
    - Metadata (Row 2 to header-2): Borders, Bold labels in Col A. Cells B-End merged.
    - Spacer (Row header-1): Left Clean.
    - Table (Row header to end): Borders, Gray Header, Auto-width.
    - custom_widths: Dict of column letters to specific widths (e.g. {'B': 42.0, 'D': 20.0})
    """
    if custom_widths is None:
        custom_widths = {}

    table_max_col = 0
    for cell in ws[header_row_num]:
        if cell.value:
            table_max_col = max(table_max_col, cell.column)
    
    if table_max_col == 0:
        table_max_col = ws.max_column
        if table_max_col == 0: table_max_col = 1 # Fallback final

    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    fill_color = "D9D9D9" 
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    max_row = ws.max_row
    
    ranges_to_remove = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= 1 and merged_range.max_row >= 1:
            ranges_to_remove.append(merged_range)
    
    for r in ranges_to_remove:
        ws.merged_cells.remove(r)
    
    if table_max_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table_max_col)

    title_cell = ws.cell(row=1, column=1)
    title_cell.fill = header_fill
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for col in range(1, table_max_col + 1):
        ws.cell(row=1, column=col).border = thin_border

    metadata_end_row = header_row_num - 2
    if metadata_end_row >= 2:
        for row_idx in range(2, metadata_end_row + 1):
            
            ranges_in_row = []
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row == row_idx and merged_range.max_row == row_idx:
                     ranges_in_row.append(merged_range)
            for r in ranges_in_row:
                ws.merged_cells.remove(r)
            
            if table_max_col > 1:
                ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=table_max_col)

            for col_idx in range(1, table_max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                if col_idx == 1:
                    cell.font = Font(bold=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

    for row in ws.iter_rows(min_row=header_row_num, max_row=max_row, min_col=1, max_col=table_max_col):
        for cell in row:
            cell.border = thin_border
            
            if cell.row == header_row_num:
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
            else:
                if isinstance(cell.value, (int, float)) or str(cell.value).strip() == "N/A":
                     cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                     cell.alignment = Alignment(horizontal='left', vertical='center')

    col_widths = {}
    
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=table_max_col):
        for cell in row:
            val = cell.value
            if val:
                length = len(str(val))
                col_letter = get_column_letter(cell.column)
                
                if cell.row == header_row_num:
                    length += 3 
                
                current_width = col_widths.get(col_letter, 0)
                if length > current_width:
                     col_widths[col_letter] = length

    for col_idx in range(1, table_max_col + 1):
        col_letter = get_column_letter(col_idx)
        width = col_widths.get(col_letter, 12) 
        
        if col_letter in custom_widths:
            adjusted_width = float(custom_widths[col_letter])
        else:
            adjusted_width = min(max(width + 2, 12), 60) 
            
        ws.column_dimensions[col_letter].width = adjusted_width
